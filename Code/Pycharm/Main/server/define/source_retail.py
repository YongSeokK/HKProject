import json
import traceback
from copy import copy
from datetime import timedelta, datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from tqdm import tqdm

##### 초기 값 (config 로 관리 할 것) #####
folder_path = 'C:\\Users\\hk_edu\\Documents\\UiPath\\Price_info\\'


##### 초기 값 (config 로 관리 할 것) #####


class KAMISDataRetail:

    ## 초기 설정
    def __init__(self, folder_path):

        self.folder_path = folder_path  # 폴더 경로

        self.weekday = datetime.today().weekday()  # 요일 확인 용
        if self.weekday == 0:
            self.day = (datetime.today() - timedelta(2)).strftime("%Y%m%d")
        else:
            self.day = (datetime.today() - timedelta(0)).strftime("%Y%m%d")
        self.read_file = '품목별_소매가격_' + self.day + '.xlsx'  # 읽어올 excel 파일
        self.path = self.folder_path + 'Result_' + self.day + '\\Claim\\'  # 파일 경로
        self.excel_file = self.path + self.read_file  # 전체 excel 파일

        # 각 각의 key 값
        self.FoodCrop = ['쌀_일반계(20kg)', '쌀_햇일반계(20kg) ', '찹쌀_일반계(1kg)', '콩_흰 콩(국산)(500g)',
                         '팥_붉은 팥(국산)(500g)', '녹두_국산(500g)', '고구마_밤(1kg)', '감자_수미(100g)']  # 식량작물
        self.Vegetables = ['배추_가을(1포기)', '배추_월동(1포기)', '배추_봄(1포기)', '배추_고랭지(1포기)',
                           '양배추_양배추(1포기)', '시금치_시금치(1kg)', '상추_적(100g)', '상추_청(100g)',
                           '얼갈이배추_얼갈이배추(1kg)', '갓_갓(1kg)', '수박_수박(1개)', '참외_참외(10개)',
                           '오이_가시계통(10개)', '오이_다다기계통(10개)', '오이_취청(10개)',
                           '호박_애호박(1개)', '호박_쥬키니(1개)', '토마토_토마토(1kg)', '딸기_딸기(100g)',
                           '무_가을(1개)', '무_월동(1개)', '무_봄(1개)', '무_고랭지(1개)',
                           '당근_무세척(1kg)', '열무_열무(1kg)', '건고추_화건(600g)', '건고추_양건(600g)',
                           '풋고추_풋고추(100g)', '풋고추_꽈리고추(100g)', '풋고추_청양고추(100g)', '붉은고추_붉은고추(100g)',
                           '양파_양파(1kg)', '파_대파(1kg)', '파_쪽파(1kg)', '생강_국산(1kg)', '고춧가루_국산(1kg)',
                           '미나리_미나리(100g)', '깻잎_깻잎(100g)', '피망_청(100g)', '파프리카_파프리카(200g)', '멜론_멜론(1개)',
                           '깐마늘(국산)_깐마늘(국산)(1kg)', '방울토마토_방울토마토(1kg)', '방울토마토_대추방울토마토(1kg)']  # 채소류
        self.SpecialCrops = ['참깨_백색(국산)(500g)', '땅콩_국산(100g)', '느타리버섯_느타리버섯(100g)', '느타리버섯_애느타리버섯(100g)',
                             '팽이버섯_팽이버섯(150g)', '새송이버섯_새송이버섯(100g)', '호두_수입(100g)', '아몬드_수입(100g)']  # 특용작물
        self.Fruit = ['사과_후지(10개)', '사과_쓰가루(10개)', '사과_홍로(10개)', '배_신고(10개)', '배_원황(10개)', '복숭아_백도(10개)',
                      '포도_캠벨얼리(1kg)', '포도_거봉(2kg)', '포도_MBA(1kg)', '포도_샤인머스켓(2kg)', '포도_레드글로브 페루(1kg)',
                      '감귤_노지(10개)', '감귤_시설(10개)', '단감_단감(10개)', '바나나_수입(100g)', '참다래_그린 뉴질랜드(10개)',
                      '참다래_국산(10개)', '파인애플_수입(1개)', '오렌지_네이블 미국(10개)', '오렌지_네이블 호주(10개)', '레몬_수입(10개)',
                      '체리_수입(100g)', '건포도_수입(100g)', '건블루베리_수입(100g)', '망고_수입(1개)']  # 과일류

        self.MarineProducts = ['고등어_생선(1마리)', '고등어_냉동(1마리)', '고등어_염장(2마리)', '꽁치_냉동(수입)(5마리)',
                               '갈치_생선(1마리)', '갈치_냉동(1마리)', '명태_냉동(1마리)', '물오징어_생선(1마리)', '물오징어_냉동(1마리)',
                               '건멸치_건멸치(100g)', '건오징어_건오징어(10마리)', '김_마른김(10장)', '김_얼구운김(10장)',
                               '건미역_건미역(100g)', '굴_굴(1kg)', '수입조기_부세수입(냉동)(1마리)', '새우젓_새우젓(1kg)',
                               '멸치액젓_멸치액젓(1kg)', '굵은소금_굵은소금(5kg)', '전복_전복(5마리)', '새우_흰다리(수입)(10마리)']  # 수산물

        self.ListDict = {'식량작물': self.FoodCrop, '채소류': self.Vegetables, '특용작물': self.SpecialCrops,
                         '과일류': self.Fruit, '수산물': self.MarineProducts}

    ## json 파일 만들기
    def Make_Json(self):
        try:
            wb = load_workbook(filename=self.excel_file)
            cnt = 0
            region_1 = ''
            Result_DF = pd.DataFrame()  # 최종 json 파일로 나갈 데이터 자료

            Result = {}
            TempList = []

            for sheet_name in tqdm(wb.sheetnames, desc='Make json file'):
                sheet = wb[sheet_name]
                str_sheet = str(sheet)
                # print('시트 명 : ' + str_sheet)

                ## 'Sheet1' 와 '품목별_소매가격' Sheet 자료는 필요 없는 자료이므로 제외
                if (str_sheet != '<Worksheet "Sheet1">') & (str_sheet != '<Worksheet "품목별_소매가격">'):
                    if cnt == 0:
                        region_1 = (str_sheet.split('"')[1]).split("_")[0]
                        region_2 = (str_sheet.split('"')[1]).split("_")[0]
                    else:
                        region_2 = (str_sheet.split('"')[1]).split("_")[0]

                        ## 지역 변경 (ex : 위에까진 '서울' 자료를 불러오다가 아래부터 '부산' 자료를 불러 올때)
                        if region_1 != region_2:
                            # print(region_1)
                            Result['Result'] = TempList
                            # print(Result)
                            with open(self.path + self.day + "_" + region_1 + '.json', 'w', encoding='utf-8') as f:
                                json.dump(Result, f, indent=4, ensure_ascii=False)
                            f.close()

                            Result = {}
                            TempList = []
                            region_1 = region_2
                    cnt += 1

                    ## 병합된 액셀 셀 처리해주는 작업
                    mcr_coord_list = [mcr.coord for mcr in sheet.merged_cells.ranges]
                    for mcr in mcr_coord_list:
                        min_col, min_row, max_col, max_row = range_boundaries(mcr)
                        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
                        top_left_cell_alignment = sheet.cell(row=min_row, column=min_col).alignment
                        top_left_cell_format = sheet.cell(row=min_row, column=min_col).number_format
                        sheet.unmerge_cells(mcr)

                        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                            for cell in row:
                                cell.value = top_left_cell_value
                                cell.alignment = copy(top_left_cell_alignment)
                                cell.number_format = copy(top_left_cell_format)

                    ## 중복발생하는 필요없는 행, 열 삭제
                    sheet.delete_rows(0, 2)
                    sheet.delete_cols(6, 13)

                    ## 작물별 List 처리 , 이중 List 구문
                    division = (str_sheet.split('"')[1]).split("_")[1]
                    # print('작물 구분 : ' + division)

                    for division in self.ListDict[division]:
                        # print(division + ' 찾기 시작')

                        traditional_cnt = 0
                        traditional_price = 0
                        market_cnt = 0
                        market_price = 0

                        for row in sheet:
                            Temp = row[0].value + "_" + row[1].value
                            if division == Temp:

                                traditional_cnt += 1
                                market_cnt += 1
                                # print(r[3].value, r[4].value)

                                if (row[3].value == '-') & (row[4].value == '-'):
                                    traditional_price = traditional_price + int((row[3].value).replace('-', '0'))
                                    market_price = market_price + int((row[4].value).replace('-', '0'))
                                    traditional_cnt -= 1
                                    market_cnt -= 1

                                elif (row[3].value == '-'):
                                    traditional_price = traditional_price + int((row[3].value).replace('-', '0'))
                                    market_price = market_price + int((row[4].value).replace(',', ''))
                                    traditional_cnt -= 1

                                elif (row[4].value == '-'):
                                    traditional_price = traditional_price + int((row[3].value).replace(',', ''))
                                    market_price = market_price + int((row[4].value).replace('-', '0'))
                                    market_cnt -= 1

                                else:
                                    traditional_price = traditional_price + int((row[3].value).replace(',', ''))
                                    market_price = market_price + int((row[4].value).replace(',', ''))

                                if traditional_price != 0:
                                    traditional_price = int(traditional_price / traditional_cnt)
                                if market_price != 0:
                                    market_price = int(market_price / market_cnt)

                        # print('종류 : ' + c + ', 전통시장 가격 : ' + str(traditional_price) + ", 마켓 가격 : " + str(market_price))
                        Temp_Dict = {}
                        Temp_Dict[division] = [traditional_price, market_price]
                        # print(Temp_Dict)
                        TempList.append(Temp_Dict)

            Result['Result'] = TempList
            # print(region_1)
            # print(Result)

            with open(self.path + self.day + "_" + region_1 + '.json', 'w', encoding='utf-8') as f:
                json.dump(Result, f, indent=4, ensure_ascii=False)
            f.close()

            return '처리 완료'

        except Exception as e:
            message = traceback.format_exc()
            return message

# My_KAMISDataRetail = KAMISDataRetail(folder_path)
# My_KAMISDataRetail.Make_Json()
